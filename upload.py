import psycopg2
from psycopg2 import Error
import argparse
import openpyxl

parser = argparse.ArgumentParser()
parser.add_argument('path', type=str, help='path to file')
args = parser.parse_args()
path = (args.path)

# wb = openpyxl.load_workbook('/Users/shirvani/Downloads/test_tasks/upload_names/dots_names.xlsm')
wb = openpyxl.load_workbook(path)
ws = wb.active


def open_db():
    """ Открывает соединение к БД Postgresql находящуюся в Docker контейнере """
    global connection
    try:
        connection = psycopg2.connect(user="postgres",
                                      password="pass",
                                      host="127.0.0.1",
                                      port="6000",
                                      database="test-encost")
        print("Соединение с PostgreSQL открыто")
    except (Exception, Error) as error:
        print("Ошибка при работе с PostgreSQL:", error)


def close_db():
    """ Закрывает соединение с БД """
    global connection
    if connection:
        cursor.close()
        connection.close()
        print("Соединение с PostgreSQL закрыто")


open_db()
name_Table = "dots_names"
cursor = connection.cursor()
cursor.execute("select exists(select * from information_schema.tables where table_name='dots_names')")

if not (cursor.fetchone()[0]):
    sqlCreateTable = ''' create table " +
                     name_Table + " (endpoint_id integer UNIQUE NOT NULL, endpoint_name varchar(128));'''
    # Create a table in PostgreSQL database
    cursor.execute(sqlCreateTable)
    connection.commit()
    print(f"Создана новая таблица {name_Table}")

query = '''INSERT INTO dots_names (endpoint_id, endpoint_name) 
                VALUES (%s, %s)
                ON CONFLICT (endpoint_id) DO UPDATE SET endpoint_name = EXCLUDED.endpoint_name;'''

first_row = ws.iter_rows(values_only=True)
print(first_row)
for row in ws.iter_rows(values_only=True):
    if type(row[0]) is int:
        try:
            cursor.execute(query, row)
            print("Данные записаны в БД")
        except (Exception) as error:
            print("Не удалось произвести запись в БД: ", error)
    elif (row[0] == None):
        pass
    else:
        print("Ключ должен быть целочисленным, ваш ключ " + "\"" + str(row[0]) + "\"")
connection.commit()

close_db()
